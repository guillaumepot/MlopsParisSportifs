"""
Get Datas from source - Matchs history
"""

"""
Libraries
"""
import requests
import pandas as pd
import zipfile
import os
from openpyxl import load_workbook

from common_variables import path_data_raw, current_season

"""
Functions
"""
def scrapMatchHistory(url, filename):
    """
    Downloads a file from the given URL and saves it with the specified filename.

    Args:
        url (str): The URL of the file to be downloaded.
        filename (str): The name of the file to be saved.

    Returns:
        dict: A dictionary with a message indicating the status of the download and save operation.
            If the operation is successful, the message will be "File downloaded and saved successfully".
            If the operation fails, the message will be "request returned with status code {status_code}".
    """
    # get file according to URL
    response = requests.get(url, filename)

    if response.status_code == 200:
        with open(filename, 'wb') as file:
            file.write(response.content)
        return {"message" : "File downloaded and saved successfully"}

    else:
        raise Exception(f"request returned with status code {response.status_code}")



def sortDatasMatchHistory(filename, path_data_raw=path_data_raw):
    """
    Sorts the data from the given file et rename files.

    Parameters:
    filename (str): The name of the file to be sorted.

    Returns:
    None
    """
    # Check the file extension
    if filename.endswith('.zip'):
        # Unzip the file
        with zipfile.ZipFile(filename, 'r') as zip_ref:
            zip_ref.extractall(path_data_raw)

    elif filename.endswith('.xlsx'):
        # Load the workbook
        wb = load_workbook(filename)
        # Get the sheet names
        sheet_names = wb.sheetnames
        # For each sheet, read it and save it as a csv file
        for sheet in sheet_names:
            df = pd.read_excel(filename, sheet_name=sheet)
            df.to_csv(os.path.join(path_data_raw, f"{sheet}.csv"), index=False)



def renameFilesMatchHistory(dictionary, path_data_raw=path_data_raw):
    """
    Renames the files in the given directory according to the given dictionary.

    Parameters:
    directory (str): The directory where the files are located.
    dictionary (dict): The dictionary with the current filenames as keys and the new filenames as values.

    Returns:
    None
    """
    for filename in os.listdir(path_data_raw):
        base, extension = os.path.splitext(filename)
        if base in dictionary:
            os.rename(os.path.join(path_data_raw, filename), os.path.join(path_data_raw, dictionary[base] + extension))



def removeOldSeasonsFromRawDatas(dictionary, path_data_raw=path_data_raw, current_season=current_season):
    """
    Remove old seasons from raw data files.
    
    Args:
        current_season (str): The current season (e.g., 2122 = 2021-2022).
        
    Returns:
        None
    """
    files = list(dictionary.values())

    current_season_years = [int("20" + current_season[:2]),
                            int("20" + current_season[2:])]
    
    print(f"All these files will be checked: {files}", f"\n Season to keep: {current_season_years}")


    for file in files:
        file_path = os.path.join(path_data_raw, file + ".csv")
        if os.path.exists(file_path):
            # Read the CSV file
            df = pd.read_csv(file_path)

            # Convert the 'Date' column to datetime
            df['Date'] = pd.to_datetime(df['Date'])

            # Filter the data to keep only the rows from the current season
            df = df[df['Date'].dt.year.isin(current_season_years)]

            # Rewrite the CSV file with the filtered data
            df.to_csv(file_path, index=False)



def scrapMatchHistoryAndSortDatas(url, filename, dictionary, league="main"):
    """
    Scrapes match history from a given URL and sorts the data.

    Parameters:
    url (str): The URL to scrape match history from.
    filename (str): The name of the file to save the scraped data.
    dictionary (dict): A dictionary containing mappings for renaming files.
    league (str, optional): The league to scrape match history from. Defaults to "main".

    Returns:
    dict: A dictionary with a success message.

    """
    scrapMatchHistory(url, filename)
    sortDatasMatchHistory(filename)
    renameFilesMatchHistory(dictionary)

    if league=='other':
        removeOldSeasonsFromRawDatas(dictionary)

    return {"message" : "Match results history datas have been downloaded and sorted successfully"}